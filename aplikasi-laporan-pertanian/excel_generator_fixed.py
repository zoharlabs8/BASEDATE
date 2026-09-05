import re, copy
from pathlib import Path
from typing import List, Dict, Any
import openpyxl

TEMPLATE_PATH_DEFAULT = Path(__file__).parent.parent / "sample" / "TAMPLATE DATA .xlsx"
SATUAN_MAP = {"bambu":"Rumpun","pisang":"Pohon","porang":"Pohon","nenas":"Pohon","nanas":"Pohon"}
def _safe(name): 
    s=re.sub(r'[:\\/?*\[\]]','-',name.strip())
    return s[:31] if len(s)>31 else s or "Sheet1"
def _satuan(k, raw=None):
    if raw: return raw
    for key,val in SATUAN_MAP.items():
        if key in k.lower(): return val
    return "Pohon"

def generate_excel(template_path, records, output_path):
    template_path=Path(template_path); output_path=Path(output_path)
    wb=openpyxl.load_workbook(template_path)
    # Keep original active ws as template master
    template_ws = wb.active
    # We will rebuild workbook correctly: first rename template_ws for record0, then copy template_ws for rest BEFORE filling
    # Step 1: create sheets with correct names by copying template_ws
    first_name=_safe(records[0].get("sheet_name") or records[0].get("nama_pemilik") or "Data-1")
    template_ws.title=first_name
    sheet_names=[first_name]
    for rec in records[1:]:
        nm=_safe(rec.get("sheet_name") or rec.get("nama_pemilik") or "Data")
        base=nm; c=1
        while nm in wb.sheetnames:
            nm=_safe(base[:28]+f"_{c}"); c+=1
        ws_copy=wb.copy_worksheet(template_ws)
        ws_copy.title=nm
        sheet_names.append(nm)
    # Step 2: fill each sheet
    for idx, rec in enumerate(records):
        ws=wb[sheet_names[idx]]
        nub=rec.get("nub","") or ""
        ws["F8"]=f": {nub}" if nub and not nub.startswith(":") else (nub if nub.startswith(":") else f": {nub}" if nub else ": ")
        ht=rec.get("hari_tanggal","") or ""
        ws["F9"]=f": {ht}" if ht and not ht.startswith(":") else (ht if ht.startswith(":") else f": {ht}" if ht else ": ")
        np=rec.get("nama_pemilik","") or ""
        ws["F10"]=f": {np}"
        ng=rec.get("nama_penggarap","") or np
        ws["F11"]=f": {ng}"
        luas=rec.get("luas_lahan","") or ""
        ws["F12"]=f": {luas} M² (Kebun)" if luas and "M" not in str(luas) else (f": {luas}" if luas else ":                  M² (Kebun)")
        komoditi_list=rec.get("komoditi",[]) or []
        S=15; E=21; JR=22; cap=E-S+1
        n=len(komoditi_list)
        # FIX: semua merged di bawah JR harus digeser karena insert_rows tidak geser merged
        merged_to_shift=[]
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row >= JR:
                merged_to_shift.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))
                ws.unmerge_cells(str(mr))
        if n>cap:
            offset = n-cap
            ws.insert_rows(JR, amount=offset)
            # re-merge dengan offset
            for r1,c1,r2,c2 in merged_to_shift:
                ws.merge_cells(start_row=r1+offset, start_column=c1, end_row=r2+offset, end_column=c2)
        else:
            # tidak insert, merge kembali tanpa offset
            for r1,c1,r2,c2 in merged_to_shift:
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        # clear old if less
        if n<cap:
            for r in range(S+n, E+1):
                for c in range(1,9):
                    ws.cell(row=r, column=c).value=None
        thin=openpyxl.styles.Side(style="thin", color="000000"); medium=openpyxl.styles.Side(style="medium", color="000000")
        for i,kom in enumerate(komoditi_list):
            row=S+i
            ws.cell(row=row,column=1).value=i+1
            ws.cell(row=row,column=1).alignment=openpyxl.styles.Alignment(horizontal="center",vertical="center")
            ws.cell(row=row,column=2).value=kom.get("nama","")
            ws.cell(row=row,column=2).alignment=openpyxl.styles.Alignment(horizontal="left",vertical="center")
            ws.cell(row=row,column=3).value=int(kom.get("kecil",0) or 0)
            ws.cell(row=row,column=4).value=int(kom.get("sedang",0) or 0)
            ws.cell(row=row,column=5).value=int(kom.get("besar",0) or 0)
            for col in (3,4,5,6):
                ws.cell(row=row,column=col).alignment=openpyxl.styles.Alignment(horizontal="center",vertical="center")
            ws.cell(row=row,column=6).value=f"=SUM(C{row}:E{row})"
            ws.cell(row=row,column=7).value=kom.get("satuan") or _satuan(kom.get("nama",""))
            ws.cell(row=row,column=7).alignment=openpyxl.styles.Alignment(horizontal="center",vertical="center")
            ws.cell(row=row,column=8).value=kom.get("keterangan") or "Tahunan"
            ws.cell(row=row,column=8).alignment=openpyxl.styles.Alignment(horizontal="left",vertical="center")
            for col in range(1,9):
                cell=ws.cell(row=row,column=col)
                cell.border=openpyxl.styles.Border(left=medium if col==1 else thin, right=medium if col==8 else thin, top=thin, bottom=thin)
                cell.font=openpyxl.styles.Font(name="Calibri",size=11,bold=False)
        jrow=JR+max(0,n-cap)
        # merge Jumlah di posisi baru (sudah di-unmerge sebelum insert)
        ws.merge_cells(start_row=jrow,start_column=1,end_row=jrow,end_column=2)
        ws.cell(row=jrow,column=1).value="Jumlah"
        ws.cell(row=jrow,column=1).alignment=openpyxl.styles.Alignment(horizontal="left",vertical="center")
        ws.cell(row=jrow,column=1).font=openpyxl.styles.Font(name="Calibri",size=11,bold=True)
        start=S; end=S+n-1 if n>0 else S
        if n==0:
            for cc in [3,4,5,6]: ws.cell(row=jrow,column=cc).value=0
        else:
            ws.cell(row=jrow,column=3).value=f"=SUM(C{start}:C{end})"
            ws.cell(row=jrow,column=4).value=f"=SUM(D{start}:D{end})"
            ws.cell(row=jrow,column=5).value=f"=SUM(E{start}:E{end})"
            ws.cell(row=jrow,column=6).value=f"=SUM(F{start}:F{end})"
        for cc in (3,4,5,6):
            ws.cell(row=jrow,column=cc).alignment=openpyxl.styles.Alignment(horizontal="center",vertical="center")
            ws.cell(row=jrow,column=cc).font=openpyxl.styles.Font(name="Calibri",size=11,bold=True)
        for col in range(1,9):
            c=ws.cell(row=jrow,column=col)
            c.border=openpyxl.styles.Border(left=medium if col==1 else thin, right=medium if col==8 else thin, top=thin, bottom=medium)
        offset=jrow-JR
        batas=rec.get("batas",{}) or {}
        for key, orig in [("utara",46),("timur",47),("selatan",48),("barat",49)]:
            r=orig+offset; val=batas.get(key,"") 
            if val: ws.cell(row=r,column=3).value=f":  {val}" if not str(val).strip().startswith(":") else val
        ws.cell(row=30+offset,column=1).value=np
    wb.save(output_path)
    return output_path
